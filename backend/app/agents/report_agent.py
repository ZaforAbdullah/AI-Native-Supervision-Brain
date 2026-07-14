from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from app.models.models import Advisor, Finding, AnalysisRun, SupervisionReport
from app.config import settings
from app.services.ai_client import get_ai_client
import logging
import uuid
import os
from datetime import datetime

logger = logging.getLogger(__name__)

REPORTS_DIR = "reports"
os.makedirs(REPORTS_DIR, exist_ok=True)


class ReportAgent:
    def __init__(self, db: Session):
        self.db = db
        self.ai = get_ai_client()

    def generate_network_report(
        self,
        analysis_run: AnalysisRun,
        advisors: List[Advisor],
        all_findings: Dict[int, List[Finding]],
        edd_notes: Dict[int, str],
    ) -> SupervisionReport:
        report_ref = f"RPT-{datetime.utcnow().strftime('%Y%m%d')}-{str(uuid.uuid4())[:6].upper()}"

        critical_advisors = [a for a in advisors if a.current_risk_grade == "critical"]
        high_advisors = [a for a in advisors if a.current_risk_grade == "high"]
        medium_advisors = [a for a in advisors if a.current_risk_grade == "medium"]

        total_findings = sum(len(f) for f in all_findings.values())
        summary = (
            f"Supervision analysis run {analysis_run.run_ref} completed on "
            f"{datetime.utcnow().strftime('%d %b %Y')}. "
            f"Analysed {len(advisors)} advisors. "
            f"Identified {total_findings} risk findings across {analysis_run.risks_identified} advisors. "
            f"Critical: {len(critical_advisors)}, High: {len(high_advisors)}, "
            f"Medium: {len(medium_advisors)}."
        )

        ai_summary_used = False
        if self.ai.available:
            ai_summary = self._generate_ai_summary(
                analysis_run, advisors, all_findings, critical_advisors, high_advisors, medium_advisors
            )
            if ai_summary:
                summary = ai_summary
                ai_summary_used = True

        pdf_path = self._generate_pdf(
            report_ref, analysis_run, advisors, all_findings, edd_notes,
            critical_advisors, high_advisors, medium_advisors, summary, ai_summary_used
        )
        excel_path = self._generate_excel(
            report_ref, analysis_run, advisors, all_findings
        )

        report = SupervisionReport(
            report_ref=report_ref,
            analysis_run_id=analysis_run.id,
            report_type="network",
            title=f"Network Risk Analysis Report — {datetime.utcnow().strftime('%B %Y')}",
            summary=summary,
            total_advisors=len(advisors),
            high_risk_count=len(high_advisors),
            critical_risk_count=len(critical_advisors),
            pdf_path=pdf_path,
            excel_path=excel_path,
        )
        self.db.add(report)
        self.db.commit()
        self.db.refresh(report)
        return report

    def _generate_ai_summary(
        self,
        analysis_run: AnalysisRun,
        advisors: List[Advisor],
        all_findings: Dict[int, List[Finding]],
        critical: List[Advisor],
        high: List[Advisor],
        medium: List[Advisor],
    ) -> Optional[str]:
        try:
            priority = sorted(critical + high, key=lambda a: a.current_risk_score, reverse=True)[:5]
            lines = []
            for a in priority:
                top_titles = [f.title for f in all_findings.get(a.id, [])[:2]]
                lines.append(
                    f"- {a.full_name} ({a.advisor_ref}, {a.firm_name}): "
                    f"{a.current_risk_grade.upper()} — " + "; ".join(top_titles)
                )
            priority_text = "\n".join(lines) if lines else "None"

            prompt = f"""You are a compliance supervision analyst writing the executive summary for a Network Risk Analysis report.

Analysis Run: {analysis_run.run_ref}
Advisors Analysed: {len(advisors)}
Risk Distribution: {len(critical)} Critical, {len(high)} High, {len(medium)} Medium, {len(advisors) - len(critical) - len(high) - len(medium)} Low
Total Findings: {analysis_run.risks_identified}

Priority Advisors (Critical/High):
{priority_text}

Write a concise (3-4 sentence) executive summary for a compliance officer audience. Focus on the overall
risk picture, any notable concentration of risk, and the priority action required. Use professional
supervision/compliance language. Do not use markdown formatting."""

            return self.ai.generate(prompt, max_tokens=400)
        except Exception as e:
            logger.warning(f"AI executive summary generation failed: {e}")
            return None

    def _generate_pdf(
        self,
        report_ref: str,
        analysis_run: AnalysisRun,
        advisors: List[Advisor],
        all_findings: Dict[int, List[Finding]],
        edd_notes: Dict[int, str],
        critical: List[Advisor],
        high: List[Advisor],
        medium: List[Advisor],
        summary: str,
        ai_summary_used: bool,
    ) -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, PageBreak
            )
            from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

            pdf_path = os.path.join(REPORTS_DIR, f"{report_ref}.pdf")
            doc = SimpleDocTemplate(pdf_path, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=18, spaceAfter=6)
            heading1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=14, textColor=colors.HexColor("#1e3a5f"))
            heading2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=11, textColor=colors.HexColor("#2563eb"))
            normal = styles["Normal"]
            small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=8)

            grade_colors = {
                "critical": colors.HexColor("#dc2626"),
                "high": colors.HexColor("#ea580c"),
                "medium": colors.HexColor("#ca8a04"),
                "low": colors.HexColor("#16a34a"),
            }

            story = []

            story.append(Paragraph("AI Supervision Brain — POC", styles["Normal"]))
            story.append(Paragraph("Network Risk Analysis Report", title_style))
            story.append(Paragraph(f"Report Reference: {report_ref}", normal))
            story.append(Paragraph(f"Analysis Run: {analysis_run.run_ref}", normal))
            story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%d %B %Y %H:%M UTC')}", normal))
            story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e5e7eb")))
            story.append(Spacer(1, 0.4*cm))

            story.append(Paragraph("Executive Summary", heading1))
            if ai_summary_used:
                story.append(Paragraph(summary, normal))
            else:
                story.append(Paragraph(
                    f"This report presents the results of an automated Network Risk Analysis conducted against "
                    f"{len(advisors)} active advisors. The analysis applied {self.db.query(type('X', (), {})).count() if False else 'configured'} "
                    f"supervision rules across five key risk datasets. "
                    f"<b>{len(critical)}</b> advisor(s) were graded Critical, <b>{len(high)}</b> High, "
                    f"and <b>{len(medium)}</b> Medium risk.",
                    normal
                ))
            story.append(Spacer(1, 0.3*cm))

            summary_data = [
                ["Risk Grade", "Count", "% of Total"],
                ["Critical", str(len(critical)), f"{len(critical)/max(len(advisors),1)*100:.1f}%"],
                ["High", str(len(high)), f"{len(high)/max(len(advisors),1)*100:.1f}%"],
                ["Medium", str(len(medium)), f"{len(medium)/max(len(advisors),1)*100:.1f}%"],
                ["Low", str(len(advisors)-len(critical)-len(high)-len(medium)),
                 f"{(len(advisors)-len(critical)-len(high)-len(medium))/max(len(advisors),1)*100:.1f}%"],
            ]
            t = Table(summary_data, colWidths=[6*cm, 3*cm, 3*cm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ]))
            story.append(t)
            story.append(Spacer(1, 0.5*cm))

            priority_advisors = critical + high
            if priority_advisors:
                story.append(Paragraph("Priority Findings — Critical & High Risk Advisors", heading1))
                story.append(PageBreak())

                for advisor in priority_advisors[:20]:
                    advisor_findings = all_findings.get(advisor.id, [])
                    grade_color = grade_colors.get(advisor.current_risk_grade, colors.grey)

                    story.append(Paragraph(
                        f'<font color="#{advisor.current_risk_grade == "critical" and "dc2626" or "ea580c"}">'
                        f'[{advisor.current_risk_grade.upper()}]</font> '
                        f'{advisor.full_name} — {advisor.advisor_ref}',
                        heading2
                    ))
                    story.append(Paragraph(f"Firm: {advisor.firm_name} | Risk Score: {advisor.current_risk_score:.1f}/10", small))
                    story.append(Spacer(1, 0.2*cm))

                    for finding in advisor_findings:
                        story.append(Paragraph(f"• <b>{finding.title}</b>", normal))
                        if finding.description:
                            story.append(Paragraph(f"  {finding.description}", small))
                        if finding.ai_analysis:
                            story.append(Paragraph(f"  <i>Analysis: {finding.ai_analysis}</i>", small))
                        story.append(Spacer(1, 0.15*cm))

                    edd = edd_notes.get(advisor.id, "")
                    if edd:
                        story.append(Paragraph("Enhanced Due Diligence Notes:", ParagraphStyle(
                            "EDDHead", parent=normal, fontName="Helvetica-Bold", fontSize=9
                        )))
                        for line in edd.split("\n")[:10]:
                            story.append(Paragraph(line, small))

                    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e5e7eb")))
                    story.append(Spacer(1, 0.3*cm))

            story.append(PageBreak())
            story.append(Paragraph("Full Advisor Risk Register", heading1))
            all_data = [["Advisor Ref", "Name", "Firm", "Risk Grade", "Risk Score", "Findings"]]
            for adv in sorted(advisors, key=lambda a: a.current_risk_score, reverse=True):
                all_data.append([
                    adv.advisor_ref,
                    adv.full_name[:25],
                    adv.firm_name[:20],
                    adv.current_risk_grade.upper(),
                    f"{adv.current_risk_score:.1f}",
                    str(len(all_findings.get(adv.id, []))),
                ])
            reg_table = Table(all_data, colWidths=[2.5*cm, 4.5*cm, 4*cm, 2.5*cm, 2*cm, 2*cm])
            reg_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e5e7eb")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("ALIGN", (3, 0), (-1, -1), "CENTER"),
            ]))
            story.append(reg_table)

            story.append(Spacer(1, 0.5*cm))
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e5e7eb")))
            story.append(Paragraph(
                "This report was generated by the AI Supervision Brain POC system. "
                "Findings are based on automated rule evaluation and AI-assisted analysis. "
                "All decisions and actions must be reviewed and approved by a qualified compliance officer.",
                ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7, textColor=colors.grey)
            ))

            doc.build(story)
            return pdf_path

        except Exception as e:
            logger.error(f"PDF generation failed: {e}")
            return ""

    def _generate_excel(
        self,
        report_ref: str,
        analysis_run: AnalysisRun,
        advisors: List[Advisor],
        all_findings: Dict[int, List[Finding]],
    ) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            xlsx_path = os.path.join(REPORTS_DIR, f"{report_ref}.xlsx")
            wb = openpyxl.Workbook()

            ws_summary = wb.active
            ws_summary.title = "Summary"

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(fill_type="solid", fgColor="1E3A5F")

            ws_summary["A1"] = "AI Supervision Brain — Network Risk Analysis Report"
            ws_summary["A1"].font = Font(bold=True, size=14, color="1E3A5F")
            ws_summary["A2"] = f"Report Ref: {report_ref}"
            ws_summary["A3"] = f"Run: {analysis_run.run_ref}"
            ws_summary["A4"] = f"Generated: {datetime.utcnow().strftime('%d %B %Y %H:%M UTC')}"

            ws_summary["A6"] = "Risk Grade"
            ws_summary["B6"] = "Count"
            ws_summary["C6"] = "Percentage"
            for cell in [ws_summary["A6"], ws_summary["B6"], ws_summary["C6"]]:
                cell.font = header_font
                cell.fill = header_fill

            grades = {"critical": 0, "high": 0, "medium": 0, "low": 0}
            for adv in advisors:
                grades[adv.current_risk_grade] = grades.get(adv.current_risk_grade, 0) + 1

            grade_fills = {
                "critical": "DC2626", "high": "EA580C", "medium": "CA8A04", "low": "16A34A"
            }
            for row, (grade, count) in enumerate(grades.items(), start=7):
                ws_summary[f"A{row}"] = grade.upper()
                ws_summary[f"B{row}"] = count
                ws_summary[f"C{row}"] = f"{count/max(len(advisors),1)*100:.1f}%"
                ws_summary[f"A{row}"].fill = PatternFill(fill_type="solid", fgColor=grade_fills[grade])
                ws_summary[f"A{row}"].font = Font(color="FFFFFF", bold=True)

            ws_advisors = wb.create_sheet("Advisor Risk Register")
            headers = ["Advisor Ref", "Full Name", "Firm", "Risk Grade", "Risk Score", "Finding Count",
                       "EFM Flag", "Last Analysed"]
            for col, h in enumerate(headers, 1):
                cell = ws_advisors.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            for row, adv in enumerate(sorted(advisors, key=lambda a: a.current_risk_score, reverse=True), start=2):
                ws_advisors.cell(row=row, column=1, value=adv.advisor_ref)
                ws_advisors.cell(row=row, column=2, value=adv.full_name)
                ws_advisors.cell(row=row, column=3, value=adv.firm_name)
                grade_cell = ws_advisors.cell(row=row, column=4, value=adv.current_risk_grade.upper())
                grade_cell.fill = PatternFill(fill_type="solid", fgColor=grade_fills.get(adv.current_risk_grade, "FFFFFF"))
                grade_cell.font = Font(color="FFFFFF", bold=True)
                ws_advisors.cell(row=row, column=5, value=round(adv.current_risk_score, 2))
                ws_advisors.cell(row=row, column=6, value=len(all_findings.get(adv.id, [])))
                ws_advisors.cell(row=row, column=7, value="YES" if adv.enhanced_financial_monitoring else "No")
                ws_advisors.cell(row=row, column=8, value=adv.last_analysed_at.strftime("%d/%m/%Y") if adv.last_analysed_at else "")

            for col in range(1, len(headers) + 1):
                ws_advisors.column_dimensions[get_column_letter(col)].auto_size = True

            ws_findings = wb.create_sheet("Findings Detail")
            f_headers = ["Advisor Ref", "Advisor Name", "Finding Type", "Risk Grade",
                         "Risk Score", "Title", "Description", "Triggered Value", "Threshold", "Requires EDD"]
            for col, h in enumerate(f_headers, 1):
                cell = ws_findings.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            frow = 2
            for adv in advisors:
                for finding in all_findings.get(adv.id, []):
                    ws_findings.cell(row=frow, column=1, value=adv.advisor_ref)
                    ws_findings.cell(row=frow, column=2, value=adv.full_name)
                    ws_findings.cell(row=frow, column=3, value=finding.finding_type)
                    gc = ws_findings.cell(row=frow, column=4, value=finding.risk_grade.upper())
                    gc.fill = PatternFill(fill_type="solid", fgColor=grade_fills.get(finding.risk_grade, "FFFFFF"))
                    gc.font = Font(color="FFFFFF", bold=True)
                    ws_findings.cell(row=frow, column=5, value=round(finding.risk_score, 2))
                    ws_findings.cell(row=frow, column=6, value=finding.title)
                    ws_findings.cell(row=frow, column=7, value=finding.description)
                    ws_findings.cell(row=frow, column=8, value=finding.triggered_value)
                    ws_findings.cell(row=frow, column=9, value=finding.threshold_value)
                    ws_findings.cell(row=frow, column=10, value="Yes" if finding.requires_edd else "No")
                    frow += 1

            wb.save(xlsx_path)
            return xlsx_path

        except Exception as e:
            logger.error(f"Excel generation failed: {e}")
            return ""
